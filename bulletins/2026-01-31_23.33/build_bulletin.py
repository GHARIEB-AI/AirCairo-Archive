import json, os, sys, subprocess, datetime

json_path = r'''D:\CODERED\LAB\Artifacts\publish_repo\AirCairo-Archive\bulletins\2026-01-31_23.33\manifest.json'''
out_dir   = r'''D:\CODERED\LAB\Artifacts\publish_repo\AirCairo-Archive\bulletins\2026-01-31_23.33'''
xlsx_path = os.path.join(out_dir, "bulletin_manifest.xlsx")
pc_html   = os.path.join(out_dir, "pc_report.html")
mob_html  = os.path.join(out_dir, "mob_report.html")

# ensure openpyxl
try:
    import openpyxl
except Exception:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

with open(json_path, "r", encoding="utf-8") as f:
    rows = json.load(f)

# Create XLSX
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "manifest"

headers = [
  "source","rel_path","file_name","ext","category","publish","reason",
  "size_mb","size_bytes","last_write","full_path"
]
ws.append(headers)

for r in rows:
    ws.append([
      r.get("source",""),
      r.get("rel_path",""),
      r.get("file_name",""),
      r.get("ext",""),
      r.get("category",""),
      str(r.get("publish","")),
      r.get("reason",""),
      r.get("size_mb",""),
      r.get("size_bytes",""),
      r.get("last_write",""),
      r.get("full_path",""),
    ])

# autosize (rough)
for col in range(1, len(headers)+1):
    maxlen = 10
    for cell in ws[get_column_letter(col)]:
        v = "" if cell.value is None else str(cell.value)
        if len(v) > maxlen: maxlen = len(v)
    ws.column_dimensions[get_column_letter(col)].width = min(maxlen + 2, 70)

# Summary sheet
ws2 = wb.create_sheet("summary")
total = len(rows)
excluded = sum(1 for r in rows if (r.get("ext","").lower()==".xlsx" and str(r.get("file_name","")).lower().startswith("run")))
publishable = sum(1 for r in rows if r.get("publish") is True)
ws2.append(["generated_at", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
ws2.append(["total_files_indexed", total])
ws2.append(["publishable_files", publishable])
ws2.append(["excluded_run_xlsx_count", excluded])

wb.save(xlsx_path)

# HTML reports
def html_escape(s):
    return (s.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
              .replace('"',"&quot;").replace("'","&#39;"))

def make_pc(rows):
    # full table with simple filter hint
    head = """<!doctype html><html><head><meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Bulletin PC Report</title>
    <style>
      body{font-family:Arial, sans-serif; margin:18px;}
      table{border-collapse:collapse; width:100%;}
      th,td{border:1px solid #ddd; padding:8px; font-size:12px; vertical-align:top;}
      th{background:#f5f5f5; position:sticky; top:0;}
      .no{color:#a00; font-weight:bold;}
      .yes{color:#060; font-weight:bold;}
    </style></head><body>
    <h1>Bulletin PC Report</h1>
    <p>Tip: Use Ctrl+F to search.</p>
    <table><thead><tr>
    """
    cols = ["source","rel_path","file_name","ext","category","publish","reason","size_mb","last_write"]
    head += "".join(f"<th>{c}</th>" for c in cols)
    head += "</tr></thead><tbody>"

    body = []
    for r in rows:
        pub = r.get("publish", False)
        pub_cls = "yes" if pub else "no"
        tds = []
        for c in cols:
            v = "" if r.get(c) is None else str(r.get(c))
            if c == "publish":
                v = "YES" if pub else "NO"
                tds.append(f'<td class="{pub_cls}">{v}</td>')
            else:
                tds.append("<td>"+html_escape(v)+"</td>")
        body.append("<tr>"+"".join(tds)+"</tr>")

    tail = "</tbody></table></body></html>"
    return head + "\n".join(body) + tail

def make_mobile(rows):
    # grouped, concise
    head = """<!doctype html><html><head><meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Bulletin Mobile Report</title>
    <style>
      body{font-family:Arial, sans-serif; margin:14px;}
      .card{border:1px solid #ddd; border-radius:10px; padding:10px; margin:10px 0;}
      .k{font-weight:bold;}
      .no{color:#a00; font-weight:bold;}
      .yes{color:#060; font-weight:bold;}
    </style></head><body>
    <h1>Bulletin Mobile Report</h1>
    """
    tail = "</body></html>"

    # prioritize excel + pc/mob reports
    prio = [r for r in rows if r.get("category") in ("excel_sheet","pc_report","mob_report")]
    other = [r for r in rows if r.get("category") not in ("excel_sheet","pc_report","mob_report")]

    def cards(items, title):
        out = [f"<h2>{title}</h2>"]
        for r in items:
            pub = r.get("publish", False)
            pub_cls = "yes" if pub else "no"
            out.append('<div class="card">')
            out.append(f'<div><span class="k">source:</span> {html_escape(str(r.get("source","")))}</div>')
            out.append(f'<div><span class="k">file:</span> {html_escape(str(r.get("file_name","")))}</div>')
            out.append(f'<div><span class="k">path:</span> {html_escape(str(r.get("rel_path","")))}</div>')
            out.append(f'<div><span class="k">category:</span> {html_escape(str(r.get("category","")))}</div>')
            out.append(f'<div><span class="k">publish:</span> <span class="{pub_cls}">{"YES" if pub else "NO"}</span></div>')
            if not pub and r.get("reason"):
                out.append(f'<div><span class="k">reason:</span> {html_escape(str(r.get("reason","")))}</div>')
            out.append('</div>')
        return "\n".join(out)

    html = head + cards(prio, "Key Items (Excel / PC / Mobile)") + cards(other[:300], "Other (first 300)") + tail
    return html

with open(pc_html, "w", encoding="utf-8") as f:
    f.write(make_pc(rows))
with open(mob_html, "w", encoding="utf-8") as f:
    f.write(make_mobile(rows))

print("OK")
print("XLSX:", xlsx_path)
print("PC :", pc_html)
print("MOB:", mob_html)
